# -*- coding: utf-8 -*-
"""
Lógica de datos del dashboard de Tiempos de reparto - Foxtrot.
La usa app.py (Flask). La persistencia (Postgres o JSON) está en storage.py.

Base incremental: cada ruta se identifica por Route ID y su TI/TML aleatorio es
determinístico por ID, así una ruta ya cargada nunca cambia de valor al actualizar.
"""

import os, json, hashlib, re
import unicodedata
from datetime import date
from io import StringIO
from urllib.error import HTTPError
from urllib.parse import urlencode
from urllib.request import Request, urlopen
import numpy as np
import pandas as pd
from openpyxl import load_workbook

import storage

# ---------------- Parámetros ajustables ----------------
TI_CENTRO  = 35
TML_CENTRO = 27.5
TI_SD, TML_SD = 7, 6
OBJ = {"tml": 30, "ti": 30, "ruta": 7, "ruta_max": 8, "alerta_h": 12, "adh": 85, "disp": 10, "disp_error": 80}

AQUI = os.path.dirname(os.path.abspath(__file__))
PLANTILLA = os.path.join(AQUI, "plantilla_dashboard.html")
RECHAZOS_API_URL = os.environ.get(
    "RECHAZOS_API_URL",
    "https://web-production-f968ec.up.railway.app/api/rechazos/diario/integracion",
)
RECHAZOS_SUCURSAL = os.environ.get("RECHAZOS_SUCURSAL", "Dolores")
RECHAZOS_SUCURSAL_ID = os.environ.get("RECHAZOS_SUCURSAL_ID", "2")
RECHAZOS_SUCURSALES_IMPORT = os.environ.get("RECHAZOS_API_SUCURSAL", "TODAS")
SATISFACCION_CSV_URL = os.environ.get(
    "SATISFACCION_CSV_URL",
    "https://docs.google.com/spreadsheets/d/e/2PACX-1vQkEVyl9kmmMf5vsi--tz5mf39u80tJoFcBzWFFLWhHuXepY5dBEqmSzXLbD0AXapFPj9DLMBqii7TA/pub?gid=0&single=true&output=csv",
)
NPS_CSV_URL = os.environ.get(
    "NPS_CSV_URL",
    "https://docs.google.com/spreadsheets/d/e/2PACX-1vQkEVyl9kmmMf5vsi--tz5mf39u80tJoFcBzWFFLWhHuXepY5dBEqmSzXLbD0AXapFPj9DLMBqii7TA/pub?gid=1806046627&single=true&output=csv",
)
DQI_CSV_URL = os.environ.get(
    "DQI_CSV_URL",
    "https://docs.google.com/spreadsheets/d/e/2PACX-1vSC5R8XlW4kETbkIDmX95n_XEVJE4JMf-NNp7wYi6mE5OAfj-EENAC9jK0-IlkN1A/pub?gid=1861746295&single=true&output=csv",
)
DPO_GKPI_URLS = [
    ("Casa Central", "Mar de Ajo", "1", os.environ.get("DPO_GKPI_MDA_URL", "https://docs.google.com/spreadsheets/d/e/2PACX-1vTRrt57z-QDSRmDblvUV6AHs_Q1og0qgW0Ec-fp1L0QjLr8R_346nhHEkKsndqka-wQUdKSc2-3PizX/pub?gid=1241670089&single=true&output=csv")),
    ("Sucursal Dolores", "Dolores", "2", os.environ.get("DPO_GKPI_DOL_URL", "https://docs.google.com/spreadsheets/d/e/2PACX-1vTRrt57z-QDSRmDblvUV6AHs_Q1og0qgW0Ec-fp1L0QjLr8R_346nhHEkKsndqka-wQUdKSc2-3PizX/pub?gid=249846040&single=true&output=csv")),
    ("Casa Central", "Mar de Ajo", "1", os.environ.get("DPO_GKPI_MDA_EXTRA_URL", "https://docs.google.com/spreadsheets/d/e/2PACX-1vTRrt57z-QDSRmDblvUV6AHs_Q1og0qgW0Ec-fp1L0QjLr8R_346nhHEkKsndqka-wQUdKSc2-3PizX/pub?gid=1883083406&single=true&output=csv")),
    ("Sucursal Dolores", "Dolores", "2", os.environ.get("DPO_GKPI_DOL_EXTRA_URL", "https://docs.google.com/spreadsheets/d/e/2PACX-1vTRrt57z-QDSRmDblvUV6AHs_Q1og0qgW0Ec-fp1L0QjLr8R_346nhHEkKsndqka-wQUdKSc2-3PizX/pub?gid=680588527&single=true&output=csv")),
]

storage.init()


def _today():
    return date.today().strftime("%Y-%m-%d")


def rng_de_ruta(rid):
    seed = int(hashlib.md5(str(rid).encode()).hexdigest()[:8], 16)
    return np.random.default_rng(seed)


def _clamp_normal(rng, center, sd, lo, hi):
    return int(round(min(max(rng.normal(center, sd), lo), hi)))


def _dispersion(pl, real):
    if pd.notna(pl) and pd.notna(real) and pl > 0:
        return round((pl - real) / pl * 100, 1)
    return None


def _num_or_none(v):
    if pd.notna(v):
        return float(v)
    return None


def _descartar_dispersion_anomala(rec):
    """Excluye dispersiones extremas causadas por una planificacion Foxtrot invalida."""
    dk, dh = rec.get("dispkm"), rec.get("disphs")
    vals = [v for v in (dk, dh) if v is not None]
    if vals and any(abs(v) > OBJ["disp_error"] for v in vals):
        rec["disp_descartada"] = True
        rec["disp_motivo"] = "Error de planificacion Foxtrot"
        rec["dispkm_original"] = dk
        rec["disphs_original"] = dh
        rec["dispkm"] = None
        rec["disphs"] = None
    return rec


def _leer_excel(f, filename=""):
    name = (filename or getattr(f, "name", "") or str(f)).lower()
    eng = "xlrd" if name.endswith(".xls") else None
    return pd.read_excel(f, engine=eng)


def _leer_excel_visitas(f, filename=""):
    name = (filename or getattr(f, "name", "") or str(f)).lower()
    eng = "xlrd" if name.endswith(".xls") else None
    sheets = pd.read_excel(f, sheet_name=None, engine=eng)
    required = {"Route ID", "Customer ID", "Visit Start Timestamp"}
    for df in sheets.values():
        if required.issubset(set(df.columns)):
            return df
    for sheet_name, df in sheets.items():
        if "visita" in str(sheet_name).lower() or "visit" in str(sheet_name).lower():
            return df
    return next(iter(sheets.values()))


def _norm_id(v):
    if pd.isna(v):
        return ""
    s = str(v).strip()
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    m = re.fullmatch(r"0*(\d+)", s)
    if m:
        return m.group(1)
    return s


def _norm_customer_id_foxtrot(v):
    s = _norm_id(v)
    if len(s) > 8 and s.isdigit():
        return str(int(s[-8:]))
    return s


def _json_items(payload):
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        for key in ("data", "items", "resultados", "results", "rows", "rechazos"):
            if isinstance(payload.get(key), list):
                return payload[key]
        if any(k in payload for k in ("fecha", "dia", "cantidad", "rechazos", "total")):
            return [payload]
    return []


def _pick_value(row, names, default=None):
    if not isinstance(row, dict):
        return default
    lower = {str(k).lower(): v for k, v in row.items()}
    for name in names:
        if name in row:
            return row[name]
        v = lower.get(name.lower())
        if v is not None:
            return v
    return default


def _to_int(v):
    if v is None or v == "":
        return 0
    try:
        return int(_to_float(v))
    except Exception:
        return 0


def _to_float(v):
    if v is None or v == "":
        return 0.0
    try:
        if isinstance(v, (int, float, np.integer, np.floating)):
            return float(v)
        s = str(v).strip().replace("\xa0", "").replace(" ", "")
        if not s:
            return 0.0
        if "," in s and "." in s:
            return float(s.replace(".", "").replace(",", ".")) if s.rfind(",") > s.rfind(".") else float(s.replace(",", ""))
        if "," in s:
            return float(s.replace(".", "").replace(",", "."))
        return float(s)
    except Exception:
        return 0.0


def _json_safe(v):
    if pd.isna(v):
        return None
    if isinstance(v, pd.Timestamp):
        return v.isoformat()
    if hasattr(v, "isoformat"):
        try:
            return v.isoformat()
        except Exception:
            pass
    if isinstance(v, (np.integer,)):
        return int(v)
    if isinstance(v, (np.floating,)):
        return float(v)
    if isinstance(v, (np.bool_,)):
        return bool(v)
    return v


def _row_raw_dict(row):
    return {str(k): _json_safe(v) for k, v in row.items()}


def _parse_fecha_ar(v):
    dt = pd.to_datetime(v, dayfirst=True, errors="coerce")
    if pd.isna(dt):
        return ""
    return dt.strftime("%Y-%m-%d")


def cargar_satisfaccion():
    rows, errors = [], []

    def norm_tipo(v):
        s = str(v or "").strip()
        low = s.lower()
        if low == "rate my delivery % de respuestas":
            return "Rate My Delivery % de respuestas"
        if low == "rmd puntaje":
            return "RMD Puntaje"
        if low == "nps gral":
            return "NPS GRAL"
        if low == "nps delivery (entrega)":
            return "NPS DELIVERY (ENTREGA)"
        if low == "% de detractores":
            return "% DE DETRACTORES"
        return s

    def leer_csv(url, fecha_col, tipo_col, resultado_col):
        req = Request(url, headers={"Accept": "text/csv"})
        with urlopen(req, timeout=20) as res:
            raw = res.read().decode("utf-8-sig", errors="replace")
        df = pd.read_csv(StringIO(raw), dtype=str).fillna("")
        out = []
        for _, r in df.iterrows():
            fecha = _parse_fecha_ar(r.get(fecha_col, ""))
            if not fecha:
                continue
            out.append({
                "anio": int(fecha[:4]),
                "mes": fecha[:7],
                "fecha": fecha,
                "tipo": norm_tipo(r.get(tipo_col, "")),
                "resultado": _to_float(str(r.get(resultado_col, "")).replace("%", "")),
            })
        return out

    try:
        rows.extend(leer_csv(SATISFACCION_CSV_URL, "Fecha", "Tipo", "Resultado"))
    except Exception:
        errors.append("No se pudo leer el CSV publicado de RMD.")
    try:
        rows.extend(leer_csv(NPS_CSV_URL, "FECHA", "TIPO", "RESULTADO"))
    except Exception:
        errors.append("No se pudo leer el CSV publicado de NPS.")
    return {"rows": rows, "error": " ".join(errors)}


def cargar_dqi():
    req = Request(DQI_CSV_URL, headers={"Accept": "text/csv"})
    try:
        with urlopen(req, timeout=30) as res:
            raw = res.read().decode("utf-8-sig", errors="replace")
        df = pd.read_csv(StringIO(raw), dtype=str).fillna("")
    except Exception:
        return {"rows": [], "error": "No se pudo leer el CSV publicado de DQI."}
    fecha_col = _pick_col(df, ["Fecha Mvto", "Fecha Movimiento", "Fecha"])
    unids_col = _pick_col(df, ["Unids", "Unidades"])
    bultos_col = _pick_col(df, ["Bultos"])
    deposito_col = _pick_col(df, ["Depósito", "Deposito"])
    articulo_col = _pick_col(df, ["Artículo", "Articulo"])
    transporte_cols = [
        c for c in [
            _pick_col(df, ["Transporte"]),
            _pick_col(df, ["Descripción Transporte", "Descripcion Transporte"]),
            _pick_col(df, ["Descripción Movimiento", "Descripcion Movimiento"]),
        ] if c is not None
    ]
    if fecha_col is None or unids_col is None:
        return {"rows": [], "error": "El CSV de DQI no trae Fecha Mvto y Unids."}
    articulos = storage.load_articulos()

    def es_entrega(row):
        if not transporte_cols:
            return True
        txt = " ".join(str(row.get(c, "")) for c in transporte_cols).lower()
        if "camion" in txt or "camión" in txt:
            return True
        if any(x in txt for x in ("roturas acarreo", "iveco", "entrega")):
            return True
        return False

    def bultos_equivalentes(row):
        bultos = _to_float(row.get(bultos_col)) if bultos_col is not None else 0.0
        unids = _to_float(row.get(unids_col))
        articulo = _norm_id(row.get(articulo_col)) if articulo_col is not None else ""
        upb = (articulos.get(articulo) or {}).get("unidades_por_bulto")
        extra = (unids / upb) if upb and upb > 0 else 0.0
        return bultos + extra

    daily = {}
    detalles = []
    for _, r in df.iterrows():
        if deposito_col is not None and _norm_id(r.get(deposito_col)) != "7":
            continue
        if not es_entrega(r):
            continue
        fecha = _parse_fecha_ar(r.get(fecha_col, ""))
        if not fecha:
            continue
        if fecha[:4] != "2026":
            continue
        bultos_eq = bultos_equivalentes(r)
        if bultos_eq <= 0:
            continue
        articulo = _norm_id(r.get(articulo_col)) if articulo_col is not None else ""
        art = articulos.get(articulo) or {}
        camion = " ".join(str(r.get(c, "")).strip() for c in transporte_cols if str(r.get(c, "")).strip())
        if not camion:
            camion = "Sin camion"
        daily[fecha] = daily.get(fecha, 0.0) + bultos_eq
        detalles.append({
            "fecha": fecha,
            "mes": fecha[:7],
            "camion": camion,
            "articulo": articulo,
            "descripcion": art.get("descripcion") or str(r.get(_pick_col(df, ["Descripción Artículo", "Descripcion Articulo"]) or "", "")).strip(),
            "bultos": round(bultos_eq, 2),
        })
    rows = [
        {"fecha": fecha, "mes": fecha[:7], "dqi": round(valor, 1)}
        for fecha, valor in sorted(daily.items())
    ]
    return {"rows": rows, "detalles": detalles, "error": ""}


def _pick_col_norm(df, candidates):
    def clean(s):
        s = "".join(c for c in unicodedata.normalize("NFD", str(s)) if unicodedata.category(c) != "Mn")
        return re.sub(r"\s+", " ", s).strip().lower()
    lookup = {clean(c): c for c in df.columns}
    for name in candidates:
        col = lookup.get(clean(name))
        if col is not None:
            return col
    for c in df.columns:
        low = clean(c)
        if any(clean(name) in low for name in candidates):
            return c
    return None


def _dpo_ok(v):
    s = str(v or "").strip().upper()
    if not s:
        return None
    return s in ("OK", "SI", "SÍ", "TRUE", "1", "A TIEMPO")


def cargar_dpo_gkpis():
    rows, errors = [], []
    for unidad, sucursal, sid_default, url in DPO_GKPI_URLS:
        try:
            req = Request(url, headers={"Accept": "text/csv", "User-Agent": "Mozilla/5.0"})
            raw = urlopen(req, timeout=30).read().decode("utf-8-sig", errors="replace")
            df = pd.read_csv(StringIO(raw), dtype=str).fillna("")
        except Exception as e:
            errors.append(f"No se pudo leer DPO {sucursal}: {e}")
            continue
        fecha_col = _pick_col_norm(df, ["Fecha"])
        camion_col = _pick_col_norm(df, ["Camion", "Camión"])
        nro_col = _pick_col_norm(df, ["N° Camion", "N Camion", "Numero Camion"])
        chofer_col = _pick_col_norm(df, ["Chofer / Responsable billetera", "Chofer"])
        ay1_col = _pick_col_norm(df, ["Ayudante1", "Ayudante 1"])
        ay2_col = _pick_col_norm(df, ["Ayudante2", "Ayudante 2"])
        up_col = _pick_col_norm(df, ["UP"])
        clientes_col = _pick_col_norm(df, ["CLIENTES", "Clientes"])
        personas_col = _pick_col_norm(df, ["PERSONAS", "Personas"])
        pallets_col = _pick_col_norm(df, ["Pallets"])
        obs_col = _pick_col_norm(df, ["Observaciones"])
        carga_col = _pick_col_norm(df, ["Cargado a tiempo?"])
        descarga_col = _pick_col_norm(df, ["Descargado a tiempo?"])
        hora_carga_col = _pick_col_norm(df, ["Hora de carga?"])
        estado_col = _pick_col_norm(df, ["Estado"])
        sid_col = _pick_col_norm(df, ["sucursal_id"])
        if fecha_col is None or camion_col is None:
            errors.append(f"DPO {sucursal}: faltan columnas Fecha/Camion.")
            continue
        for idx, r in df.iterrows():
            fecha = _parse_fecha_ar(r.get(fecha_col, ""))
            camion = str(r.get(camion_col, "") or "").strip()
            if not fecha or not camion:
                continue
            if fecha[:4] != "2026":
                continue
            sid = str(r.get(sid_col, "") or sid_default).strip() if sid_col else sid_default
            suc = "Mar de Ajo" if sid == "1" else "Dolores"
            nro = str(r.get(nro_col, "") or "").strip() if nro_col else ""
            estado = str(r.get(estado_col, "") or "").strip() if estado_col else ""
            rows.append({
                "key": f"{fecha}|{sid}|{nro or camion}|{idx}",
                "fecha": fecha,
                "mes": fecha[:7],
                "anio": int(fecha[:4]),
                "unidad": "Casa Central" if sid == "1" else "Sucursal Dolores",
                "suc": suc,
                "sucursal_id": sid,
                "camion": camion,
                "nro_camion": nro or camion,
                "chofer": str(r.get(chofer_col, "") or "").strip() if chofer_col else "",
                "ayudante1": str(r.get(ay1_col, "") or "").strip() if ay1_col else "",
                "ayudante2": str(r.get(ay2_col, "") or "").strip() if ay2_col else "",
                "up": _to_float(r.get(up_col)) if up_col else None,
                "clientes": _to_float(r.get(clientes_col)) if clientes_col else None,
                "personas": _to_float(r.get(personas_col)) if personas_col else None,
                "pallets": _to_float(r.get(pallets_col)) if pallets_col else None,
                "cargado_ok": _dpo_ok(r.get(carga_col)) if carga_col else None,
                "descargado_ok": _dpo_ok(r.get(descarga_col)) if descarga_col else None,
                "hora_carga": str(r.get(hora_carga_col, "") or "").strip() if hora_carga_col else "",
                "estado": estado,
                "recarga": bool(estado),
                "observaciones": str(r.get(obs_col, "") or "").strip() if obs_col else "",
            })
    rows = sorted(rows, key=lambda r: (r["fecha"], r["suc"], r["nro_camion"], r["chofer"]))
    return {"rows": rows, "error": " ".join(errors)}


def dqi_objetivo_bultos_mes():
    cfg = storage.load_settings().get("dqi_objetivo_bultos_mes") or {}
    val = _to_float(cfg.get("valor"))
    return val if val > 0 else 1


def guardar_dqi_objetivo(valor):
    val = _to_float(valor)
    if val <= 0:
        raise ValueError("El objetivo mensual DQI debe ser mayor a cero.")
    return storage.save_setting("dqi_objetivo_bultos_mes", {"valor": val})


def _norm_rechazo(row):
    fecha = str(_pick_value(row, ("fecha", "dia", "date", "Fecha", "Día"), "") or "")[:10]
    cantidad = _to_int(_pick_value(row, ("pedidos_rechazo", "rechazo_pedidos", "rechazos", "cantidad", "total", "count", "valor"), 0))
    motivo = str(_pick_value(row, ("motivo", "causa", "tipo", "descripcion", "descripción", "evento", "feriado"), "") or "")
    suc = str(_pick_value(row, ("sucursal", "Sucursal"), RECHAZOS_SUCURSAL) or RECHAZOS_SUCURSAL)
    sid = str(_pick_value(row, ("sucursal_id", "sucursalId", "id_sucursal"), RECHAZOS_SUCURSAL_ID) or RECHAZOS_SUCURSAL_ID)
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", fecha):
        return None
    key = f"{fecha}|{suc}"
    return {
        "key": key,
        "fecha": fecha,
        "mes": fecha[:7],
        "sucursal": suc,
        "sucursal_id": sid,
        "rechazos": cantidad,
        "motivo": motivo,
        "pedidos_pdv_atendidos": _to_int(_pick_value(row, ("pedidos_pdv_atendidos", "pedidos", "pdv_unicos"), 0)),
        "pdv_unicos": _to_int(_pick_value(row, ("pdv_unicos",), 0)),
        "nds": _to_float(_pick_value(row, ("nds",), 0)),
        "bultos": _to_float(_pick_value(row, ("bultos",), 0)),
        "rechazo_bultos": _to_float(_pick_value(row, ("rechazo_bultos", "bultos_rechazo"), 0)),
        "rechazo_bultos_total": _to_float(_pick_value(row, ("rechazo_bultos_total", "bultos_rechazo"), 0)),
        "pct_rechazo_bultos": _to_float(_pick_value(row, ("pct_rechazo_bultos",), 0)),
        "hl": _to_float(_pick_value(row, ("hl",), 0)),
        "rechazo_hl": _to_float(_pick_value(row, ("rechazo_hl", "hl_rechazo"), 0)),
        "rechazo_hl_total": _to_float(_pick_value(row, ("rechazo_hl_total", "hl_rechazo"), 0)),
        "pct_rechazo_hl": _to_float(_pick_value(row, ("pct_rechazo_hl",), 0)),
        "pallets": _to_float(_pick_value(row, ("pallets",), 0)),
        "rechazo_pallets": _to_float(_pick_value(row, ("rechazo_pallets", "pallets_rechazo"), 0)),
        "pct_rechazo_pallets": _to_float(_pick_value(row, ("pct_rechazo_pallets",), 0)),
        "salidas": _to_int(_pick_value(row, ("salidas",), 0)),
        "pct_rechazo_pedidos": _to_float(_pick_value(row, ("pct_rechazo_pedidos", "pct_rechazo", "porcentaje"), 0)),
        "pico": str(_pick_value(row, ("pico",), "")).lower() == "true",
        "feriado": str(_pick_value(row, ("feriado",), "") or ""),
        "evento": str(_pick_value(row, ("evento",), "") or ""),
    }


def _norm_rechazo_detalle(row):
    fecha = str(_pick_value(row, ("fecha", "dia", "date"), "") or "")[:10]
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", fecha):
        return None
    chofer = str(_pick_value(row, ("chofer", "driver", "repartidor"), "Sin chofer") or "Sin chofer").strip() or "Sin chofer"
    motivo = str(_pick_value(row, ("motivo", "causa"), "Sin motivo") or "Sin motivo").strip() or "Sin motivo"
    sector = str(_pick_value(row, ("sector",), "Sin sector") or "Sin sector").strip() or "Sin sector"
    suc = str(_pick_value(row, ("sucursal", "Sucursal"), RECHAZOS_SUCURSAL) or RECHAZOS_SUCURSAL)
    rec = {
        "fecha": fecha,
        "mes": fecha[:7],
        "sucursal": suc,
        "chofer": chofer,
        "chofer_codigo": str(_pick_value(row, ("chofer_codigo", "codigo_chofer"), "") or ""),
        "sector": sector,
        "motivo": motivo,
        "pedidos_rechazo": _to_int(_pick_value(row, ("pedidos_rechazo", "rechazos"), 0)),
        "ocurrencias": _to_int(_pick_value(row, ("ocurrencias", "cantidad", "total"), 0)),
        "bultos_rechazo": _to_float(_pick_value(row, ("bultos_rechazo", "rechazo_bultos"), 0)),
        "hl_rechazo": _to_float(_pick_value(row, ("hl_rechazo", "rechazo_hl"), 0)),
        "pallets_rechazo": _to_float(_pick_value(row, ("pallets_rechazo", "rechazo_pallets"), 0)),
    }
    key_parts = [fecha, suc, rec["chofer_codigo"] or chofer, sector, motivo]
    return "|".join(str(x).replace("|", "/") for x in key_parts), rec


def importar_rechazos(desde=None, hasta=None):
    desde = desde or "2026-01-01"
    hasta = hasta or _today()
    params = {"desde": desde, "hasta": hasta, "sucursal": RECHAZOS_SUCURSALES_IMPORT}
    if not RECHAZOS_API_URL.endswith("/integracion"):
        params["formato"] = "csv"
    url = RECHAZOS_API_URL + "?" + urlencode(params)
    req = Request(url, headers={"Accept": "text/csv, application/json"})
    try:
        with urlopen(req, timeout=30) as res:
            ctype = res.headers.get("Content-Type", "")
            raw = res.read().decode("utf-8")
    except HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")[:500]
        raise ValueError(f"El endpoint respondió HTTP {e.code}. URL: {url}. Respuesta: {body}") from e
    if "csv" in ctype.lower() or url.endswith("formato=csv"):
        return guardar_rechazos_csv(raw, desde, hasta, url)
    if "json" in ctype.lower():
        payload = json.loads(raw)
        return guardar_rechazos_payload(payload, desde, hasta, url)
    raise ValueError(f"El endpoint no devolvió CSV ni JSON. Content-Type: {ctype or 'sin Content-Type'}")


def _fetch_rechazos(url):
    req = Request(url, headers={"Accept": "text/csv, application/json"})
    try:
        with urlopen(req, timeout=30) as res:
            return res.read().decode("utf-8"), res.headers.get("Content-Type", ""), url
    except HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")[:500]
        raise ValueError(f"El endpoint respondio HTTP {e.code}. URL: {url}. Respuesta: {body}") from e


def importar_rechazos(desde=None, hasta=None):
    desde = desde or "2026-01-01"
    hasta = hasta or _today()
    params = {"desde": desde, "hasta": hasta, "sucursal": RECHAZOS_SUCURSALES_IMPORT}
    url = RECHAZOS_API_URL + "?" + urlencode(params)
    try:
        raw, ctype, final_url = _fetch_rechazos(url)
        if "csv" in ctype.lower() or final_url.endswith("formato=csv"):
            return guardar_rechazos_csv(raw, desde, hasta, final_url)
        if "json" in ctype.lower():
            return guardar_rechazos_payload(json.loads(raw), desde, hasta, final_url)
    except ValueError as first_error:
        if "HTTP 404" not in str(first_error) or not RECHAZOS_API_URL.endswith("/diario/integracion"):
            raise
        base = RECHAZOS_API_URL.rsplit("/diario/integracion", 1)[0]
        resumen_url = base + "/diario/resumen?" + urlencode(params)
        detalle_url = base + "/diario/detalle?" + urlencode(params)
        try:
            resumen_raw, _, _ = _fetch_rechazos(resumen_url)
            detalle_raw, _, _ = _fetch_rechazos(detalle_url)
        except ValueError as second_error:
            raise ValueError(
                "La API de rechazos no esta disponible en la URL configurada. "
                f"Probo integracion, resumen y detalle. Ultimo error: {second_error}. "
                "Verifica que la app origen tenga registrado/deployado app.routes.rechazos "
                "o configura RECHAZOS_API_URL con la URL correcta."
            ) from second_error
        resumen_payload = json.loads(resumen_raw)
        detalle_payload = json.loads(detalle_raw)
        payload = {
            "resumen_diario": resumen_payload.get("datos", resumen_payload),
            "detalle_diario": detalle_payload.get("datos", detalle_payload),
        }
        return guardar_rechazos_payload(payload, desde, hasta, resumen_url)
    raise ValueError(f"El endpoint no devolvio CSV ni JSON. URL: {url}")


def guardar_rechazos_csv(raw, desde="", hasta="", origen="archivo"):
    df = pd.read_csv(StringIO(raw))
    payload = df.fillna("").to_dict(orient="records")
    return guardar_rechazos_payload(payload, desde, hasta, origen)


def _norm_header(v):
    s = unicodedata.normalize("NFKD", str(v or "").replace("\n", " ").strip())
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", s).upper()


def guardar_rechazos_excel(file_obj, desde="", hasta="", origen="archivo"):
    if hasattr(file_obj, "seek"):
        file_obj.seek(0)
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb["BASE"] if "BASE" in wb.sheetnames else wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    header = None
    for row in rows:
        vals = [_norm_header(v) for v in row]
        if "FECHA" in vals and "BULTOS" in vals and ("UNIDAD PAQUETE" in vals or "UNIDAD DE MEDIDA" in vals):
            header = vals
            break
    if not header:
        raise ValueError("No se encontro una hoja de rechazos con columnas FECHA, BULTOS y UNIDAD PAQUETE.")
    idx = {name: i for i, name in enumerate(header) if name}

    def val(row, name, default=0):
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else default

    recs, det, docs_por_fecha, docs_detalle = {}, {}, {}, {}
    for nrow, row in enumerate(rows, start=1):
        fecha = _parse_fecha_ar(val(row, "FECHA", ""))
        if not fecha:
            continue
        bultos = _to_float(val(row, "BULTOS"))
        bultos_rech = _to_float(val(row, "BULTOS RECHAZADOS"))
        hl = _to_float(val(row, "UNIDAD PAQUETE", val(row, "UNIDAD DE MEDIDA")))
        hl_rech = _to_float(val(row, "UNIDAD PAQUETE RECHAZADO", val(row, "UNIDAD DE MEDIDA RECHAZADO")))
        pallets = _to_float(val(row, "UNIDAD DE MEDIDA"))
        pallets_rech = _to_float(val(row, "UNIDAD DE MEDIDA RECHAZADO"))
        doc = str(val(row, "DETALLE DOCUMENTO", "") or val(row, "NUMERO", "") or nrow).strip()
        rechazo_flag = str(val(row, "RECHAZO", "")).strip().upper()
        es_rechazo = bultos_rech > 0 or hl_rech > 0 or rechazo_flag not in ("", "0", "NO")

        suc = RECHAZOS_SUCURSAL
        key_rec = f"{fecha}|{suc}"
        rec = recs.setdefault(key_rec, {
            "key": key_rec, "fecha": fecha, "mes": fecha[:7], "sucursal": suc, "sucursal_id": RECHAZOS_SUCURSAL_ID,
            "rechazos": 0, "motivo": "", "pedidos_pdv_atendidos": 0, "pdv_unicos": 0, "nds": 0,
            "bultos": 0.0, "rechazo_bultos": 0.0, "rechazo_bultos_total": 0.0, "pct_rechazo_bultos": 0.0,
            "hl": 0.0, "rechazo_hl": 0.0, "rechazo_hl_total": 0.0, "pct_rechazo_hl": 0.0,
            "pallets": 0.0, "rechazo_pallets": 0.0, "pct_rechazo_pallets": 0.0,
            "salidas": 0, "pct_rechazo_pedidos": 0.0, "pico": False, "feriado": "", "evento": "",
        })
        rec["bultos"] += bultos
        rec["rechazo_bultos"] += bultos_rech
        rec["rechazo_bultos_total"] += bultos_rech
        rec["hl"] += hl
        rec["rechazo_hl"] += hl_rech
        rec["rechazo_hl_total"] += hl_rech
        rec["pallets"] += pallets
        rec["rechazo_pallets"] += pallets_rech
        if es_rechazo:
            docs_por_fecha.setdefault(key_rec, set()).add(doc)
            chofer = str(val(row, "DESCRIPCION CHOFER", val(row, "DESCRIPCION DETALLDA CHOFER", "Sin chofer")) or "Sin chofer").strip()
            motivo = str(val(row, "MOTIVO DE RECHAZO", val(row, "DESCRIPCION DETALLADA MOTIVO", "Sin motivo")) or "Sin motivo").strip()
            sector = str(val(row, "DESCRIPCION RUTA", val(row, "RUTA", "Sin sector")) or "Sin sector").strip()
            key = "|".join(x.replace("|", "/") for x in [fecha, suc, chofer, sector, motivo])
            d = det.setdefault(key, {"fecha": fecha, "mes": fecha[:7], "sucursal": RECHAZOS_SUCURSAL, "chofer": chofer,
                                     "chofer_codigo": str(val(row, "CHOFER", "") or ""), "sector": sector, "motivo": motivo,
                                     "pedidos_rechazo": 0, "ocurrencias": 0, "bultos_rechazo": 0.0,
                                     "hl_rechazo": 0.0, "pallets_rechazo": 0.0})
            docs_detalle.setdefault(key, set()).add(doc)
            d["ocurrencias"] += 1
            d["bultos_rechazo"] += bultos_rech
            d["hl_rechazo"] += hl_rech
            d["pallets_rechazo"] += pallets_rech

    for key_rec, rec in recs.items():
        rec["rechazos"] = len(docs_por_fecha.get(key_rec, set()))
        rec["pct_rechazo_bultos"] = (rec["rechazo_bultos"] / rec["bultos"] * 100) if rec["bultos"] else 0.0
        rec["pct_rechazo_hl"] = (rec["rechazo_hl"] / rec["hl"] * 100) if rec["hl"] else 0.0
        rec["pct_rechazo_pallets"] = (rec["rechazo_pallets"] / rec["pallets"] * 100) if rec["pallets"] else 0.0
    for key, item in det.items():
        item["pedidos_rechazo"] = len(docs_detalle.get(key, set()))

    guardados = storage.upsert_rechazos(recs)
    detalle_guardados = storage.upsert_rechazos_detalle(det) if det else 0
    return {"desde": desde, "hasta": hasta, "url": origen, "recibidos": len(recs), "guardados": guardados, "detalle_guardados": detalle_guardados}


def guardar_rechazos_payload(payload, desde="", hasta="", origen="archivo"):
    recs = {}
    resumen = payload.get("resumen_diario") if isinstance(payload, dict) else None
    detalle = payload.get("detalle_diario") if isinstance(payload, dict) else None
    for row in (resumen if resumen is not None else _json_items(payload)):
        rec = _norm_rechazo(row)
        if not rec:
            continue
        key = rec["key"]
        if key not in recs:
            recs[key] = rec
        else:
            recs[key]["rechazos"] += rec["rechazos"]
            recs[key]["pedidos_pdv_atendidos"] += rec.get("pedidos_pdv_atendidos", 0)
            recs[key]["pdv_unicos"] += rec.get("pdv_unicos", 0)
            recs[key]["bultos"] += rec.get("bultos", 0)
            recs[key]["rechazo_bultos"] += rec.get("rechazo_bultos", 0)
            recs[key]["rechazo_bultos_total"] += rec.get("rechazo_bultos_total", 0)
            recs[key]["hl"] += rec.get("hl", 0)
            recs[key]["rechazo_hl"] += rec.get("rechazo_hl", 0)
            recs[key]["rechazo_hl_total"] += rec.get("rechazo_hl_total", 0)
            recs[key]["pallets"] += rec.get("pallets", 0)
            recs[key]["rechazo_pallets"] += rec.get("rechazo_pallets", 0)
            if rec["motivo"] and not recs[key].get("motivo"):
                recs[key]["motivo"] = rec["motivo"]
    guardados = storage.upsert_rechazos(recs)
    det_recs = {}
    if detalle is not None:
        for row in detalle:
            item = _norm_rechazo_detalle(row)
            if item:
                key, rec = item
                det_recs[key] = rec
    detalle_guardados = storage.upsert_rechazos_detalle(det_recs) if det_recs else 0
    return {"desde": desde, "hasta": hasta, "url": origen, "recibidos": len(recs), "guardados": guardados, "detalle_guardados": detalle_guardados}


def _time_to_min(h, m="0"):
    return int(h) * 60 + int(m)


def parse_horario_entrega(texto):
    """Convierte '09:00 A 13:00 Y DE 17:00 A 21:00' en rangos en minutos."""
    if pd.isna(texto):
        return []
    s = str(texto).upper().strip()
    if not s or s in {"0", "0.00%", "NAN"}:
        return []
    s = (s.replace("HS", "").replace("HRS", "").replace("HORAS", "")
           .replace("–", " A ").replace("-", " A ").replace("A.", "A"))
    pairs = re.findall(r"(\d{1,2})(?::(\d{2}))?\s*(?:A|/|HASTA)\s*(\d{1,2})(?::(\d{2}))?", s)
    rangos = []
    for h1, m1, h2, m2 in pairs:
        ini = _time_to_min(h1, m1 or "0")
        fin = _time_to_min(h2, m2 or "0")
        if 0 <= ini < 24 * 60 and 0 < fin <= 24 * 60 and ini != fin:
            rangos.append({"ini": ini, "fin": fin})
    return rangos


def _en_ventana(ts, ventanas):
    if pd.isna(ts) or not ventanas:
        return None
    minuto = int(ts.hour) * 60 + int(ts.minute)
    for v in ventanas:
        ini, fin = v["ini"], v["fin"]
        if ini <= fin and ini <= minuto <= fin:
            return True
        if ini > fin and (minuto >= ini or minuto <= fin):
            return True
    return False


def _fmt_minuto(m):
    return f"{int(m) // 60:02d}:{int(m) % 60:02d}"


def _fmt_ventanas(ventanas):
    return " / ".join(f"{_fmt_minuto(v['ini'])}-{_fmt_minuto(v['fin'])}" for v in ventanas)


def procesar_clientes(clientes_file):
    c = pd.read_csv(clientes_file, sep=";", dtype=str, encoding="cp1252")
    out = {}
    for _, r in c.iterrows():
        cliente = _norm_id(r.get("Cliente"))
        if not cliente:
            continue
        horario = r.get("Horario de entrega")
        ventanas = parse_horario_entrega(horario)
        out[cliente] = {
            "cliente": cliente,
            "sucursal": _norm_id(r.get("Sucursal")),
            "razon_social": "" if pd.isna(r.get("Razon social")) else str(r.get("Razon social")).strip(),
            "nombre": "" if pd.isna(r.get("Nombre de fantasia")) else str(r.get("Nombre de fantasia")).strip(),
            "horario_entrega": "" if pd.isna(horario) else str(horario).strip(),
            "ventanas": ventanas,
        }
    return out


def actualizar_clientes(clientes_file):
    clientes = procesar_clientes(clientes_file)
    return storage.replace_clientes(clientes)


def procesar_articulos(articulos_file):
    try:
        df = pd.read_csv(articulos_file, dtype=str, sep=None, engine="python", encoding="utf-8-sig", encoding_errors="replace").fillna("")
    except UnicodeDecodeError:
        if hasattr(articulos_file, "seek"):
            articulos_file.seek(0)
        df = pd.read_csv(articulos_file, dtype=str, sep=None, engine="python", encoding="cp1252", encoding_errors="replace").fillna("")
    art_col = _pick_col(df, ["Artículo", "Articulo", "Codigo", "Código", "SKU"])
    desc_col = _pick_col(df, ["Descripción Artículo", "Descripcion Articulo", "Descripcion", "Descripción"])
    upb_col = _pick_col(df, [
        "Unidades por bulto", "Unidades x bulto", "Unid x bulto", "UxB",
        "Unidades/Bulto", "Unidades por caja", "Factor", "Contenido",
    ])
    if art_col is None or upb_col is None:
        raise ValueError("El archivo de articulos debe tener articulo y unidades por bulto.")
    out = {}
    for _, r in df.iterrows():
        articulo = _norm_id(r.get(art_col))
        upb = _to_float(r.get(upb_col))
        if not articulo or upb <= 0:
            continue
        out[articulo] = {
            "articulo": articulo,
            "descripcion": str(r.get(desc_col, "")).strip() if desc_col is not None else "",
            "unidades_por_bulto": upb,
        }
    return out


def actualizar_articulos(articulos_file):
    articulos = procesar_articulos(articulos_file)
    return storage.replace_articulos(articulos)


def _pick_col(df, candidates):
    lookup = {str(c).strip().lower(): c for c in df.columns}
    for name in candidates:
        col = lookup.get(name.lower())
        if col is not None:
            return col
    for c in df.columns:
        low = str(c).lower()
        if any(name.lower() in low for name in candidates):
            return c
    return None


def _leer_csv_visitas(csv_files):
    frames = []
    for f, name in csv_files:
        try:
            if hasattr(f, "seek"):
                f.seek(0)
            n = (name or getattr(f, "name", "") or "").lower()
            if n.endswith((".xls", ".xlsx")):
                c = _leer_excel_visitas(f, n)
            else:
                c = pd.read_csv(f)
        except Exception:
            continue
        if "Route ID" not in c.columns:
            continue
        c = c.copy()
        c["click"] = pd.to_datetime(c.get("Driver Click Timestamp"), errors="coerce")
        c["vs"] = pd.to_datetime(c.get("Visit Start Timestamp"), errors="coerce")
        c["visend"] = c["vs"] + pd.to_timedelta(c["Visit Duration Seconds"], "s") \
            if "Visit Duration Seconds" in c.columns else c["vs"]
        frames.append(c)
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


def _attempt_key(row, idx):
    parts = [
        row.get("Route ID"),
        row.get("Waypoint ID"),
        row.get("Customer ID"),
        row.get("Visit Start Timestamp"),
        row.get("Driver Click Timestamp"),
        idx,
    ]
    return "|".join("" if pd.isna(p) else str(p) for p in parts)


def procesar_attempts(csv_files):
    visitas = _leer_csv_visitas(csv_files)
    if visitas.empty:
        return {}
    out = {}
    for idx, row in visitas.iterrows():
        rec = _row_raw_dict(row)
        key = _attempt_key(row, idx)
        rec["attempt_key"] = key
        out[key] = rec
    return out


def _mapa_correccion(csv_files):
    c = _leer_csv_visitas(csv_files)
    if c.empty:
        return {}
    lv = c.groupby("Route ID").agg(u1=("click", "max"), u2=("visend", "max"))
    lv["f"] = lv["u1"].fillna(lv["u2"])
    return lv["f"].to_dict()


def _mapa_ontime(csv_files):
    visitas = _leer_csv_visitas(csv_files)
    clientes = storage.load_clientes()
    if visitas.empty:
        return {}
    cli_col = _pick_col(visitas, ["Customer ID", "Customer Id", "Customer", "Client ID", "Cliente"])
    if cli_col is None:
        return {}
    ts = visitas["click"].fillna(visitas["vs"])
    visitas = visitas.assign(cliente=visitas[cli_col].map(_norm_customer_id_foxtrot), paso=ts)
    stats = {}
    for rid, grp in visitas.groupby("Route ID"):
        total = ontime = fuera = sin_ventana = 0
        fuera_clientes = []
        clientes_con_ventana = {}
        clientes_sin_ventana = {}
        for _, v in grp.iterrows():
            cid = v["cliente"]
            if not cid:
                continue
            total += 1
            cliente = clientes.get(cid) or {}
            nombre = cliente.get("nombre") or cliente.get("razon_social") or str(v.get("Customer Name", ""))
            ventanas = cliente.get("ventanas", [])
            cliente_ref = {"cliente": cid, "nombre": nombre}
            if ventanas:
                clientes_con_ventana[cid] = cliente_ref
            else:
                cliente_ref["motivo"] = "sin ventana cargada" if cliente else "no encontrado en base de clientes"
                clientes_sin_ventana[cid] = cliente_ref
            ok = _en_ventana(v["paso"], ventanas)
            if ok is True:
                ontime += 1
            elif ok is False:
                fuera += 1
                fuera_clientes.append({
                    "cliente": cid,
                    "nombre": nombre,
                    "visita": v["paso"].strftime("%H:%M") if pd.notna(v["paso"]) else "",
                    "ventana": _fmt_ventanas(ventanas),
                })
            else:
                sin_ventana += 1
        if total:
            evaluables = ontime + fuera
            stats[str(rid)] = {
                "pdv_total": int(total),
                "pdv_ontime": int(ontime),
                "pdv_fuera_ontime": int(fuera),
                "pdv_sin_ventana": int(sin_ventana),
                "ontime_pct": round(100 * ontime / evaluables, 1) if evaluables else None,
                "clientes_fuera_ontime": fuera_clientes[:50],
                "clientes_con_ventana": list(clientes_con_ventana.values()),
                "clientes_sin_ventana": list(clientes_sin_ventana.values()),
            }
    return stats


def procesar_export(xls_file, xls_name="", csv_files=None):
    """Devuelve dict rid -> registro, calculado desde el export."""
    csv_files = csv_files or []
    x = _leer_excel(xls_file, xls_name)
    hl_col = _pick_col(x, ["HL", "HLS", "Hectolitros", "Hectolitro", "Hectoliter", "Hectoliters", "Volume HL", "Delivered HL", "Planned HL", "Actual HL", "Volumen HL", "Volumen Hectolitros"])
    bultos_col = _pick_col(x, ["Bultos", "Bultos Despachados", "Cases", "Delivered Cases", "Planned Cases", "Actual Cases", "Packages", "Unidades Paquete", "Unidad Paquete"])
    salidas_col = _pick_col(x, ["Salidas", "Stops", "Stops Count", "Customers", "Deliveries", "Pedidos", "PDV"])
    camion_col = _pick_col(x, ["Camion", "Camión", "Truck", "Vehicle", "Vehicle Name", "Vehicle ID", "Plate", "License Plate", "Patente", "Transporte", "Descripcion Transporte", "Descripción Transporte"])
    x["fox_ini"] = pd.to_datetime(x["Driver Marked Route Start Timestamp"], errors="coerce")
    x["fox_fin"] = pd.to_datetime(x["Driver Marked Route End Timestamp"], errors="coerce")
    x["suc"] = x["DC Name"].str.replace(" - del Palacio S.A.", "", regex=False)
    valid = x["fox_ini"].notna() & x["fox_fin"].notna()
    same_day = x["fox_ini"].dt.date == x["fox_fin"].dt.date
    x["raw_h"] = (x["fox_fin"] - x["fox_ini"]).dt.total_seconds() / 3600
    fmap = _mapa_correccion(csv_files)
    omap = _mapa_ontime(csv_files)

    x["fin_final"] = x["fox_fin"]; x["usable"] = False
    for i in x[valid].index:
        if same_day[i]:
            x.at[i, "usable"] = True
        else:
            nf = fmap.get(x.at[i, "Route ID"])
            if nf is not None and pd.notna(nf) and nf.date() == x.at[i, "fox_ini"].date() \
               and 0 < (nf - x.at[i, "fox_ini"]).total_seconds() / 60 <= 14 * 60:
                x.at[i, "fin_final"] = nf; x.at[i, "usable"] = True
    x["dur_h"] = (x["fin_final"] - x["fox_ini"]).dt.total_seconds() / 3600
    x.loc[(x["dur_h"] <= 0) | (x["dur_h"] > 14), "usable"] = False

    out = {}
    for i in x[valid].index:
        r = x.loc[i]; rid = str(r["Route ID"]); usable = bool(r["usable"])
        ah = r["dur_h"] if usable else r["raw_h"]
        rec = {"rid": rid, "suc": r["suc"], "chofer": r["Driver Name"],
               "mes": r["fox_ini"].strftime("%Y-%m"), "fecha": r["fox_ini"].strftime("%Y-%m-%d"),
               "anio": r["fox_ini"].strftime("%Y"),
               "inicio_foxtrot": r["fox_ini"].strftime("%H:%M") if pd.notna(r["fox_ini"]) else "",
               "fin_foxtrot": r["fin_final"].strftime("%H:%M") if pd.notna(r["fin_final"]) else "",
               "camion": str(r.get(camion_col, "")).strip() if camion_col is not None and str(r.get(camion_col, "")).strip() else "Sin camion",
               "usable": usable, "alerta": bool(ah > OBJ["alerta_h"]),
               "hl": _to_float(r.get(hl_col)) if hl_col is not None else 0.0,
               "bultos": _to_float(r.get(bultos_col)) if bultos_col is not None else 0.0,
               "salidas": _to_int(r.get(salidas_col)) if salidas_col is not None else 0,
               "raw_foxtrot": _row_raw_dict(r)}
        if rid in omap:
            rec.update(omap[rid])
        if usable:
            g = rng_de_ruta(rid)
            km_plan = _num_or_none(r.get("Planned Foxtrot Driving Meters"))
            km_real = _num_or_none(r.get("Total Driven Meters"))
            hs_plan = _num_or_none(r.get("Planned Foxtrot Driving Seconds"))
            hs_real = _num_or_none(r.get("Total Driven Seconds"))
            rec.update({"ti": _clamp_normal(g, TI_CENTRO, TI_SD, 25, 45),
                        "tml": _clamp_normal(g, TML_CENTRO, TML_SD, 20, 45),
                        "horas": round(r["dur_h"], 3),
                        "adhsec": round(r["Sequence Adherence"] * 100, 1) if pd.notna(r.get("Sequence Adherence")) else None,
                        "adhcli": round(r["Driver Click Score"] * 100, 1) if pd.notna(r.get("Driver Click Score")) else None,
                        "disp_km_plan": km_plan,
                        "disp_km_real": km_real,
                        "disp_hs_plan": hs_plan,
                        "disp_hs_real": hs_real,
                        "dispkm": _dispersion(km_plan, km_real),
                        "disphs": _dispersion(hs_plan, hs_real)})
            _descartar_dispersion_anomala(rec)
        out[rid] = rec
    return out


def _data_desde_base(base):
    rutas = sorted((_descartar_dispersion_anomala(dict(r)) for r in base.values()), key=lambda r: (r["fecha"], r["suc"], r["chofer"]))
    rechazos_base = storage.load_rechazos()
    if rutas and not rechazos_base:
        try:
            fechas = sorted(r["fecha"] for r in rutas if r.get("fecha"))
            if fechas:
                importar_rechazos(fechas[0], fechas[-1])
                rechazos_base = storage.load_rechazos()
        except Exception:
            rechazos_base = {}
    rechazos = sorted(rechazos_base.values(), key=lambda r: r["fecha"])
    rechazos_detalle = sorted(storage.load_rechazos_detalle().values(), key=lambda r: (r["fecha"], r.get("chofer", ""), r.get("motivo", "")))
    return {"rutas": rutas,
            "rechazos": rechazos,
            "rechazos_detalle": rechazos_detalle,
            "satisfaccion": cargar_satisfaccion(),
            "dqi": cargar_dqi(),
            "dpo": cargar_dpo_gkpis(),
            "settings": {"dqi_objetivo_bultos_mes": dqi_objetivo_bultos_mes()},
            "choferes": sorted({r["chofer"] for r in rutas}),
            "sucursales": sorted({r["suc"] for r in rutas}),
            "meses": sorted({r["mes"] for r in rutas}),
            "obj": OBJ}


def actualizar(xls_file, xls_name, csv_files=None, reset=False):
    """Procesa el export y agrega a la base SOLO las rutas nuevas. Devuelve stats."""
    if reset:
        storage.reset()
    previas = len(storage.load_all())
    nuevos = procesar_export(xls_file, xls_name, csv_files)
    attempts = procesar_attempts(csv_files) if csv_files else {}
    actualiza_existentes = True
    agregadas = storage.upsert_all(nuevos)
    attempts_guardados = storage.upsert_attempts(attempts) if attempts else 0
    base = storage.load_all()
    us = [r for r in base.values() if r.get("usable")]
    tml = [r["tml"] for r in us]; ti = [r["ti"] for r in us]
    return {"previas": previas, "agregadas": agregadas, "actualiza_existentes": actualiza_existentes, "procesadas": len(nuevos), "attempts_guardados": attempts_guardados, "total": len(base),
            "validas": len(us), "sin_cierre": len(base) - len(us),
            "tml_prom": round(float(np.mean(tml)), 1) if tml else None,
            "tml_cumpl": round(100 * float(np.mean([v <= 30 for v in tml]))) if tml else None,
            "ti_prom": round(float(np.mean(ti)), 1) if ti else None,
            "ti_cumpl": round(100 * float(np.mean([v <= 30 for v in ti]))) if ti else None}


def render_dashboard():
    data = _data_desde_base(storage.load_all())
    html = open(PLANTILLA, encoding="utf-8").read()
    return html.replace("__DATA__", json.dumps(data, ensure_ascii=False))


def hay_datos():
    return len(storage.load_all()) > 0
