# -*- coding: utf-8 -*-
"""
pipeline_pedidos.py
-------------------
Lee el export "reporte-pedidos" (hoja "ag-grid", una fila por comprobante) y
lo normaliza a una lista de dicts listos para persistir en Postgres.

El foco del análisis es FECHA/HORA DE ALTA = momento en que se cargó el pedido
(la "franja horaria" de ingreso). Todo lo demás son ejes para cruzar:
localidad (parseada de RUTA DE DIST.), vendedor, canal (ORIGEN), día, categoría
de comprobante y $ (TOTAL). Los bultos se dejan en None: este export no los trae
(CANT. PAQUETES viene en 0). Cuando haya fuente de bultos se completa 'bultos'.
"""

from datetime import datetime
import openpyxl

# ---- Configuración de negocio -------------------------------------------------

# Franjas horarias de ingreso. (inicio_incluido, fin_excluido, etiqueta)
FRANJAS = [
    "Antes 07",
    "07-08",
    "08-09",
    "09-10",
    "10-11",
    "11:00-11:15",
    "11:15-11:30",
    "11:30-11:45",
    "11:45-12:00",
    "12:00-12:15",
    "12:15-12:30",
    "12:30-12:45",
    "12:45-13:00",
    "13:00-13:15",
    "13:15-13:30",
    "13:30-13:45",
    "13:45-14:00",
    "14:00-14:15",
    "14:15-14:30",
    "14:30-14:45",
    "14:45-15:00",
    "15+",
]
DIAS = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]

# Mapeo de TIPO DE COMPROBANTE -> categoría operativa
CATEGORIA_COMPROBANTE = {
    "FACTURA": "VENTA",
    "FACTURA PRESUPUESTO": "PRESUPUESTO",
    "NOTA DE CREDITO": "NC",
    "DEVOLUCION PRESUPUESTO": "DEVOLUCION",
    "REMITO RESPALDO": "REMITO",
}


# ---- Helpers ------------------------------------------------------------------

def _parse_dt(v):
    """Devuelve datetime o None. Acepta datetime nativo o 'dd/mm/YYYY HH:MM:SS'."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _split_cod_nombre(v):
    """'52 - VILLALBA CAROLINA' -> ('52', 'VILLALBA CAROLINA'). Tolerante a None."""
    if v is None:
        return (None, None)
    s = str(v).strip()
    if " - " in s:
        cod, nombre = s.split(" - ", 1)
        return (cod.strip(), nombre.strip())
    return (None, s)


def _localidad_de_ruta(ruta):
    """'523 - DOLORES' -> 'DOLORES'. None/'' -> 'SIN RUTA'."""
    if ruta is None or str(ruta).strip() == "":
        return "SIN RUTA"
    _cod, nombre = _split_cod_nombre(ruta)
    return (nombre or "SIN RUTA").upper()


def _si_no(v):
    """'SÍ'/'SI' -> True, 'NO' -> False, otro -> False."""
    if v is None:
        return False
    return str(v).strip().upper() in ("SÍ", "SI", "S", "TRUE", "1")


def _franja(dt):
    hour = dt.hour
    if hour < 7:
        return "Antes 07"
    if hour < 11:
        return f"{hour:02d}-{hour + 1:02d}"
    if hour < 15:
        start_min = (dt.minute // 15) * 15
        end_hour = hour
        end_min = start_min + 15
        if end_min == 60:
            end_hour += 1
            end_min = 0
        return f"{hour:02d}:{start_min:02d}-{end_hour:02d}:{end_min:02d}"
    return "15+"


# ---- Parser principal ---------------------------------------------------------

def parse_pedidos(source, sheet_name="ag-grid"):
    """
    source: ruta a .xlsx o file-like (BytesIO del upload).
    Devuelve (records, meta) donde records es list[dict] y meta trae contadores.
    """
    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

    it = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(it)]
    idx = {h: i for i, h in enumerate(headers)}

    def g(row, name):
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else None

    records = []
    saltadas = 0
    for row in it:
        if row is None or all(c is None for c in row):
            continue
        nro = g(row, "NÚMERO PEDIDO")
        if nro is None:
            saltadas += 1
            continue

        alta = _parse_dt(g(row, "FECHA/HORA DE ALTA"))
        if alta is None:
            saltadas += 1
            continue

        vend_cod, vend_nom = _split_cod_nombre(g(row, "VENDEDOR DEL PEDIDO"))
        ruta = g(row, "RUTA DE DIST.")
        tipo_comp = (str(g(row, "TIPO DE COMPROBANTE")).strip()
                     if g(row, "TIPO DE COMPROBANTE") is not None else None)
        anulado = _si_no(g(row, "ANULADO"))

        try:
            total = float(g(row, "TOTAL") or 0)
        except (TypeError, ValueError):
            total = 0.0

        rec = {
            "nro_pedido": int(nro),
            "fecha_alta": alta.isoformat(),
            "anio": alta.year,
            "mes": alta.month,
            "fecha": alta.strftime("%Y-%m-%d"),
            "hora": alta.hour,
            "minuto": alta.minute,
            "franja": _franja(alta),
            "dia_semana": alta.weekday(),          # 0=Lun … 6=Dom
            "dia_nombre": DIAS[alta.weekday()],
            "sucursal": g(row, "SUCURSAL"),
            "origen": g(row, "ORIGEN"),            # BEES / PREVENTA / DISP.MOVIL
            "vendedor_cod": vend_cod,
            "vendedor": vend_nom,
            "cliente": g(row, "CLIENTE"),
            "ruta": (str(ruta).strip() if ruta is not None else None),
            "localidad": _localidad_de_ruta(ruta),
            "tipo_comprobante": tipo_comp,
            "categoria": CATEGORIA_COMPROBANTE.get(tipo_comp, "OTRO"),
            "anulado": anulado,
            "facturado": _si_no(g(row, "FACTURADO")),
            "fecha_entrega": (lambda d: d.strftime("%Y-%m-%d") if d else None)(
                _parse_dt(g(row, "FECHA ENTREGA"))),
            "total": round(total, 2),
            "bultos": None,   # <- hueco para completar cuando haya fuente de bultos
        }
        records.append(rec)

    wb.close()
    meta = {"leidas": len(records), "saltadas": saltadas}
    return records, meta


if __name__ == "__main__":
    import sys, json
    path = sys.argv[1] if len(sys.argv) > 1 else "reporte-pedidos__5_.xlsx"
    recs, meta = parse_pedidos(path)
    print("META:", meta)
    print("EJEMPLO:", json.dumps(recs[0], ensure_ascii=False, indent=2))
